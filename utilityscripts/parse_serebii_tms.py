from bs4 import BeautifulSoup
from openpyxl import load_workbook

if __name__ == '__main__':
	with open('gen9tms.htm') as f:
		wb = load_workbook("TM Data.xlsx")
		ws = wb.active
		soup = BeautifulSoup(f, 'html.parser')
		table = soup.find(class_="data-table")
		table_rows = table.find_all('tr')
		for tag in table_rows[1:]:
			tm_num = int(tag.find("td", class_="cell-num").string)
			tm_name = tag.find("td", class_="cell-name").find("a", class_="ent-name").string
			tm_type = tag.find("td", class_="cell-icon").find("a", class_="type-icon").string
			tm_power = tag.find_all("td", class_="cell-num")[1].string
			gems_required = 1
			# Some weird encoding stuff going on here, we actually are checking for — and ∞ 	
			if (tm_power != 'â€”'):
				tm_power_num = int(tm_power)
				if tm_power_num > 70: gems_required = 2
				if tm_power_num > 100: gems_required = 3
			print(f"{tm_name}: {tm_power}")
			
			ws.cell(row=tm_num+1, column=1).value = tm_num
			ws.cell(row=tm_num+1, column=2).value = tm_name
			ws.cell(row=tm_num+1, column=3).value = f"{tm_type} Gem"
			ws.cell(row=tm_num+1, column=4).value = gems_required
		wb.save("New TM Data.xlsx")

