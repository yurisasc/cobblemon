# Make sure to pip install openpyxl, also have all spreadsheets in current working directory
# Also make sure to pip install requests
import json
import requests
from openpyxl import load_workbook


def stuff():
    berry_data_wb = load_workbook(filename="Cobblemon Berry Data.xlsx", data_only=True)
    berry_dict = {}
    berry_file_name = ""
    for row in berry_data_wb.active.iter_rows(2):
        berry_num = int(row[0].value)
        berry_name = row[1].value
        berry_prefix = berry_name[:-6].lower()
        berry_file_name = berry_name.lower().replace(" ", "_") + ".json"
        baseYields = row[4].value.split("-")
        berry_dict["baseYield"] = {
            "min": int(baseYields[0]),
            "max": int(baseYields[1])
        }
        biomeTags = row[2].value.split(", ")
        berry_dict["preferredBiomeTags"] = [f"cobblemon:is_{x[1:].lower()}" for x in biomeTags]
        baseGrowthTimes = int(row[7].value)
        growthVar = int(row[8].value) // 2
        berry_dict["growthTime"] = {
            "min": baseGrowthTimes - growthVar,
            "max": baseGrowthTimes + growthVar
        }
        refreshRate = int(row[9].value)
        refreshRateVar = int(row[10].value) // 2
        berry_dict["refreshRate"] = {
            "min": refreshRate - refreshRateVar,
            "max": refreshRate + refreshRateVar
        }
        fav_mulches = [i.lower() for i in row[3].value.split(", ")]
        berry_dict["favoriteMulches"] = fav_mulches
        berry_dict["growthFactors"] = []
        berry_dict["randomizedGrowthPoints"] = True
        spawn_type = row[11].value
        if spawn_type == "PREFERRED_BIOME":
            berry_dict["spawnConditions"] = [
                {
                    "variant": "cobblemon:preferred_biome",
                    "minGroveSize": 3,
                    "maxGroveSize": 5
                }
            ]
        elif spawn_type == "ALL_BIOME":
            berry_dict["spawnConditions"] = [
                {
                    "variant": "cobblemon:all_biome",
                    "minGroveSize": 3,
                    "maxGroveSize": 5
                }
            ]
        else:
            berry_dict["spawnConditions"] = []
        berry_dict["growthPoints"] = get_growth_points(row[0].row)
        berry_dict["stageOnePositioning"] = get_stage_one_y_pos(row[0].row)
        berry_dict["mutations"] = get_mutations(berry_prefix)
        berry_dict["sproutShape"] = f"{row[19].value.lower()}-sprout"
        berry_dict["matureShape"] = f"{row[20].value.lower()}-mature"
        spicyVal = int(row[14].value)
        dryVal = int(row[15].value)
        sweetVal = int(row[16].value)
        bitterVal = int(row[17].value)
        sourVal = int(row[18].value)
        flavorDict = {}
        if spicyVal > 0:
            flavorDict["SPICY"] = spicyVal
        if dryVal > 0:
            flavorDict["DRY"] = dryVal
        if sweetVal > 0:
            flavorDict["SWEET"] = sweetVal
        if bitterVal > 0:
            flavorDict["BITTER"] = bitterVal
        if sourVal > 0:
            flavorDict["SOUR"] = sourVal
        berry_dict["flavors"] = flavorDict
        if row[12].value != "N/A":
            berry_dict["weight"] = float(row[12].value)
        else:
            berry_dict["weight"] = 1.0
        berry_dict["tintIndexes"] = []
        berry_dict["flowerModel"] = f"cobblemon:flower.geo"
        berry_dict["flowerTexture"] = f"cobblemon:flower"
        berry_dict["fruitModel"] = f"cobblemon:{berry_prefix}_berry.geo"
        berry_dict["fruitTexture"] = f"cobblemon:{berry_prefix}"
        berry_dict["boneMealChance"] = int(row[21].value) / 100.0
        with open(f"berries/{berry_file_name}", "w+") as file:
            file.write(json.dumps(berry_dict, indent=2))


def get_growth_points(rowNum):
    berry_placements_wb = load_workbook(filename="Berry Placements.xlsx", data_only=True)
    row = berry_placements_wb.active[rowNum + 1]
    numSpots = 0
    for i in range(0, 10):
        if row[11 + i * 6].value is None:
            break
        else:
            numSpots = numSpots + 1
    result = [
        {
            "position": {
                "x": float(row[6 + i * 6].value),
                "y": float(row[7 + i * 6].value),
                "z": float(row[8 + i * 6].value)
            },
            "rotation": {
                "x": float(row[9 + i * 6].value),
                "y": float(row[10 + i * 6].value),
                "z": float(row[11 + i * 6].value)
            }
        } for i in range(0, numSpots)
    ]
    return result

def get_stage_one_y_pos(rowNum):
    berry_placements_wb = load_workbook(filename="Berry Placements.xlsx", data_only=True)
    row = berry_placements_wb.active[rowNum + 1]
    return  {
            "position": {
                "x": float(row[66].value),
                "y": float(row[67].value),
                "z": float(row[68].value)
            },
            "rotation": {
                "x": float(row[69].value),
                "y": float(row[70].value),
                "z": float(row[71].value)
            }
        }

def get_mutations(berry_name):
    berry_mutations_wb = load_workbook(filename="Berry Mutations (v2).xlsx", data_only=True)
    result = {}
    for row in berry_mutations_wb.active.iter_rows(2):
        add_mutation = False
        berryIndex = 0
        for i, cell in enumerate(row[0:3]):
            if cell.value.lower() == berry_name:
                add_mutation = True
                berryIndex = i
        if add_mutation:
            other_berry = "cobblemon:" + (row[2].value.lower() if berryIndex == 0 else row[0].value.lower()) + "_berry"
            result_berry = "cobblemon:" + (row[4].value.lower()) + "_berry"
            result[other_berry] = result_berry
    return result


def get_tints(berry_num, berry_name):
    preceding_zero = "0" if berry_num < 10 else ""
    url = f"https://gitlab.com/Apion/cobblemon-assets/-/raw/tint36/blockbench/berry_trees/{preceding_zero}{berry_num}_{berry_name}/{berry_name}_tints.json?ref_type=heads"
    req = requests.get(url)
    result = {}
    if req.status_code == 200:
        tint_list = req.json()
        for i, x in enumerate(tint_list):
            result[str(i)] = x
    else:
        print(f"Couldn't get tintIndexes for {berry_name} berry")
    return result


if __name__ == "__main__":
    stuff()
